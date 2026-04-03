"""Match tender against historical reference database (reference_db/ JSONs).

Two independent evaluation axes for every reference:
  1. Amount gate: strict yes/no (reference value >= tender minimum)
  2. Technical match: 0-100% (how closely the type of work matches)

Output: full comparison table of ALL references with both columns,
short explanation for each, longer explanation for top 3.
"""
import base64
import json
import logging
import os
import re
import time
from datetime import datetime
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_env_path = Path(__file__).parent.parent / "briefer" / ".env"
if _env_path.exists():
    load_dotenv(_env_path, override=True)

logger = logging.getLogger(__name__)

MATCHER_DIR = Path(__file__).parent
DB_DIR = MATCHER_DIR / "reference_db"

# ---------------------------------------------------------------------------
# Normalization helpers
# ---------------------------------------------------------------------------

OPERATOR_ALIASES = {
    "pge dystrybucja": "PGE", "pge": "PGE",
    "tauron dystrybucja": "TAURON", "tauron": "TAURON",
    "enea operator": "ENEA", "enea": "ENEA",
    "energa-operator": "ENERGA", "energa operator": "ENERGA", "energa": "ENERGA",
    "polskie sieci elektroenergetyczne": "PSE", "pse": "PSE",
    "aldesa": "ALDESA", "altis": "ALTIS",
    "emca volt": "EMCA", "emca": "EMCA",
}


def _normalize_operator(raw: str) -> str:
    if not raw:
        return ""
    low = raw.lower().strip()
    for alias, canonical in sorted(OPERATOR_ALIASES.items(), key=lambda x: -len(x[0])):
        if alias in low:
            return canonical
    return raw.split()[0].upper()


def _normalize_voltage(raw: str | None) -> str:
    if not raw:
        return ""
    m = re.search(r"(\d+)\s*k[vV]", str(raw))
    return f"{m.group(1)} kV" if m else str(raw).strip()


def _normalize_work_type(raw: str | None) -> str:
    if not raw:
        return ""
    low = raw.lower().strip()
    if "przebudowa" in low: return "przebudowa"
    if "rekonstrukcja" in low: return "przebudowa"
    if "modernizacja" in low: return "modernizacja"
    if "budowa" in low: return "budowa"
    if "wymiana" in low and "izolacj" in low: return "wymiana izolacji"
    if "wymiana" in low and "opgw" in low: return "wymiana OPGW"
    if "opgw" in low: return "OPGW"
    return low


# ---------------------------------------------------------------------------
# Reference DB
# ---------------------------------------------------------------------------

def load_reference_db() -> list[dict]:
    """Load all reference JSONs from reference_db/."""
    if not DB_DIR.exists():
        logger.warning(f"No reference_db/ directory at {DB_DIR}")
        return []
    refs = []
    for json_path in sorted(DB_DIR.glob("*.json")):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            data["_db_file"] = json_path.name
            refs.append(data)
        except (json.JSONDecodeError, OSError) as e:
            logger.warning(f"Failed to load {json_path.name}: {e}")
    logger.info(f"Loaded {len(refs)} references from reference_db/")
    return refs


# ---------------------------------------------------------------------------
# Tender requirements extraction
# ---------------------------------------------------------------------------

def _extract_min_required_amount(tender_data: dict) -> float | None:
    requirements = tender_data.get("wymagania_doswiadczenie", [])
    if not requirements:
        return None
    combined = " ".join(str(r) for r in requirements).lower()
    amount_patterns = [
        r'(\d{1,3}(?:\s\d{3})+(?:[,\.]\d+)?)\s*(?:zł|pln|złotych|netto|brutto)',
        r'(\d{1,3}(?:\.\d{3})+(?:,\d+)?)\s*(?:zł|pln|złotych|netto|brutto)',
        r'(\d+(?:[,\.]\d+)?)\s*mln\b',
        r'(\d{6,}(?:[,\.]\d+)?)\s*(?:zł|pln|złotych|netto|brutto)',
    ]
    amounts = []
    for pattern in amount_patterns:
        for m in re.finditer(pattern, combined):
            raw = m.group(1)
            if "mln" in combined[m.start():m.end() + 5]:
                val = float(raw.replace(",", ".").replace(" ", "")) * 1_000_000
            else:
                cleaned = raw.replace(" ", "").replace(".", "")
                if "," in cleaned:
                    parts = cleaned.rsplit(",", 1)
                    val = float(parts[0] + "." + parts[1])
                else:
                    val = float(cleaned)
            amounts.append(val)
    if amounts:
        result = max(amounts)
        logger.info(f"Extracted min required reference amount: {result:,.0f} PLN")
        return result
    return None


def _extract_tender_requirements(tender_data: dict) -> dict:
    reqs = {
        "operator": "",
        "voltage": "",
        "work_type": "",
        "min_amount": None,
        "min_poles": None,
        "required_scope": [],
        "desired_scope": [],
        "has_opgw": False,
    }
    reqs["operator"] = _normalize_operator(tender_data.get("zamawiajacy", ""))
    nazwa = (tender_data.get("nazwa", "") or "").lower()
    zakres = (tender_data.get("zakres_prac", "") or "").lower()
    combined = nazwa + " " + zakres
    vm = re.search(r"(\d+)\s*kv", combined)
    if vm:
        reqs["voltage"] = f"{vm.group(1)} kV"
    for wt in ["przebudowa", "modernizacja", "budowa", "wymiana izolacji", "wymiana opgw"]:
        if wt in combined:
            reqs["work_type"] = wt
            break
    if "opgw" in combined or "światłowod" in combined:
        reqs["has_opgw"] = True

    # Hard scope from wymagania_doswiadczenie
    requirements = tender_data.get("wymagania_doswiadczenie", [])
    req_text = " ".join(str(r) for r in requirements).lower()
    SCOPE_KW = {
        "słupy": ["słup", "stanowisk"],
        "fundamenty": ["fundament"],
        "przewody fazowe": ["przewod", "przewód fazow"],
        "izolatory": ["izolator", "izolacj"],
        "opgw": ["opgw", "światłowod", "odgromow"],
        "demontaż": ["demontaż", "rozbiórk"],
        "uziemienia": ["uziemien"],
    }
    for element, keywords in SCOPE_KW.items():
        if any(kw in req_text for kw in keywords):
            reqs["required_scope"].append(element)

    reqs["min_amount"] = _extract_min_required_amount(tender_data)

    pole_patterns = [
        r"co najmniej\s+(\d+)\s+s\S*up",
        r"minimum\s+(\d+)\s+s\S*up",
        r"min\.?\s+(\d+)\s+s\S*up",
        r"(\d+)\s+s\S*up\S*\s+kratow",
    ]
    for pp in pole_patterns:
        pm = re.search(pp, req_text)
        if pm:
            reqs["min_poles"] = int(pm.group(1))
            break

    # Soft scope from zakres_prac + zestawienie
    SCOPE_ELEMENTS = [
        "słupy kratowe", "słupy", "fundamenty", "przewody fazowe", "przewody",
        "izolatory", "łańcuchy izolatorowe", "opgw", "światłowód", "światłowod",
        "uziemienia", "demontaż", "tłumiki drgań", "tłumiki", "tabliczki",
        "drogi dojazdowe", "mostki", "odstępniki", "zawiesia", "kable",
        "przewody kablowe", "spawy", "dokumentacja", "projekt",
    ]
    scope_text = combined
    for item in (tender_data.get("zestawienie_materialow") or []):
        scope_text += " " + (item.get("kategoria", "") or "").lower()
        for sub in (item.get("podkategorie") or []):
            scope_text += " " + sub.lower()
    for elem in SCOPE_ELEMENTS:
        if elem in scope_text and elem not in reqs["required_scope"]:
            reqs["desired_scope"].append(elem)

    return reqs


# ---------------------------------------------------------------------------
# Scope matching helpers
# ---------------------------------------------------------------------------

def _ref_has_scope_element(ref_scope: list[str], element: str) -> bool:
    element_lower = element.lower()
    for s in ref_scope:
        s_lower = s.lower()
        if element_lower in s_lower or s_lower in element_lower:
            return True
    aliases = {
        "słupy": ["słup", "stanowisk", "konstrukcj", "kratow"],
        "fundamenty": ["fundament"],
        "przewody fazowe": ["przewod fazow", "przewód fazow", "afl", "ac 300"],
        "izolatory": ["izolator", "izolacj", "łańcuch"],
        "opgw": ["opgw", "światłowod", "odgromow skojarz"],
        "demontaż": ["demontaż", "rozbiórk"],
        "uziemienia": ["uziemien", "zaziemien"],
    }
    for alias in aliases.get(element_lower, []):
        for s in ref_scope:
            if alias in s.lower():
                return True
    return False


NEW_BUILD_SIGNALS = [
    "montaż nowych", "dostawa i montaż", "zabudowa", "budowa fundament",
    "budowa linii", "demontaż", "montaż słup", "montaż przewod",
    "wykonanie fundament",
]
MAINTENANCE_SIGNALS = [
    "naprawa", "wzmocnienie", "podwyższenie", "regulacja zwis",
    "oględziny", "przegląd", "malowanie", "nivelacja",
]


def _classify_work_depth(ref: dict) -> str:
    scope = [s.lower() for s in (ref.get("scope_elements") or [])]
    summary = (ref.get("scope_summary") or "").lower()
    combined = " ".join(scope) + " " + summary
    new_count = sum(1 for sig in NEW_BUILD_SIGNALS if sig in combined)
    maint_count = sum(1 for sig in MAINTENANCE_SIGNALS if sig in combined)
    if new_count >= 2 and new_count > maint_count:
        return "new_build"
    elif maint_count >= 2 and maint_count > new_count:
        return "maintenance"
    elif new_count >= 1:
        return "hybrid"
    return "maintenance"


# ---------------------------------------------------------------------------
# Gate 1: Amount (strict yes/no)
# ---------------------------------------------------------------------------

def _check_amount_gate(ref: dict, min_amount: float | None) -> tuple[bool, str]:
    """Returns (passes, reason)."""
    if min_amount is None:
        return True, "brak wymagania kwotowego"
    ref_amount = ref.get("contract_value_pln_netto")
    if ref_amount is None:
        return False, "brak danych o wartości kontraktu"
    if ref_amount >= min_amount:
        return True, f"{ref_amount:,.0f} ≥ {min_amount:,.0f} PLN"
    return False, f"{ref_amount:,.0f} < {min_amount:,.0f} PLN"


# ---------------------------------------------------------------------------
# Gate 2: Technical match (0-100%)
# ---------------------------------------------------------------------------

def _score_technical_match(ref: dict, reqs: dict) -> tuple[int, str]:
    """Score how technically close the reference work is to what the tender requires.

    Returns (score 0-100, short_explanation).

    Scoring breakdown (100 points total):
      - Work type match:           25 pts (przebudowa vs modernizacja vs wymiana OPGW)
      - Work depth (new vs maint): 15 pts (new construction vs repair/maintenance)
      - Voltage match:             10 pts
      - Required scope elements:   20 pts (proportional to how many matched)
      - Additional scope overlap:  15 pts (proportional to desired elements matched)
      - OPGW match:                 5 pts
      - Pole count match:          10 pts (has enough poles if required)
    """
    ref_scope = [e for e in (ref.get("scope_elements") or [])]
    ref_work_type = _normalize_work_type(ref.get("work_type"))
    ref_voltage = _normalize_voltage(ref.get("voltage"))
    ref_poles = ref.get("pole_count")
    work_depth = _classify_work_depth(ref)

    score = 0
    parts = []

    # 1. Work type (25 pts)
    tender_wt = reqs["work_type"].lower() if reqs["work_type"] else ""
    ref_wt = ref_work_type.lower()
    if tender_wt and ref_wt:
        if ref_wt == tender_wt:
            score += 25
            parts.append(f"typ prac: {ref_wt} (identyczny)")
        elif tender_wt == "przebudowa" and ref_wt == "budowa":
            score += 20  # budowa is close to przebudowa
            parts.append(f"typ prac: {ref_wt} (zbliżony)")
        elif tender_wt == "przebudowa" and ref_wt == "modernizacja":
            score += 8  # modernizacja is partial
            parts.append(f"typ prac: {ref_wt} (częściowo zbliżony)")
        else:
            parts.append(f"typ prac: {ref_wt} (niezgodny z {tender_wt})")

    # 2. Work depth (15 pts)
    if tender_wt in ("przebudowa", "budowa"):
        if work_depth == "new_build":
            score += 15
            parts.append("nowa budowa/przebudowa")
        elif work_depth == "hybrid":
            score += 8
            parts.append("prace mieszane (budowa + utrzymanie)")
        else:
            parts.append("prace utrzymaniowe/naprawcze")
    else:
        score += 15  # Not relevant for non-rebuild tenders

    # 3. Voltage (10 pts)
    if reqs["voltage"] and ref_voltage:
        req_kv_m = re.search(r"(\d+)", reqs["voltage"])
        ref_kv_m = re.search(r"(\d+)", ref_voltage)
        if req_kv_m and ref_kv_m:
            req_kv = int(req_kv_m.group(1))
            ref_kv = int(ref_kv_m.group(1))
            if ref_kv == req_kv:
                score += 10
                parts.append(f"napięcie {ref_voltage} (identyczne)")
            elif ref_kv > req_kv:
                score += 8
                parts.append(f"napięcie {ref_voltage} (wyższe)")
            else:
                parts.append(f"napięcie {ref_voltage} (niższe niż {reqs['voltage']})")

    # 4. Required scope (20 pts)
    if reqs["required_scope"]:
        matched = [e for e in reqs["required_scope"] if _ref_has_scope_element(ref_scope, e)]
        ratio = len(matched) / len(reqs["required_scope"])
        pts = int(20 * ratio)
        score += pts
        if matched:
            parts.append(f"wymagany zakres ({len(matched)}/{len(reqs['required_scope'])}): {', '.join(matched)}")
        if len(matched) < len(reqs["required_scope"]):
            missing = [e for e in reqs["required_scope"] if e not in matched]
            parts.append(f"BRAK: {', '.join(missing)}")
    else:
        score += 20  # No specific scope required

    # 5. Additional scope overlap (15 pts)
    if reqs["desired_scope"]:
        matched = [e for e in reqs["desired_scope"] if _ref_has_scope_element(ref_scope, e)]
        ratio = len(matched) / len(reqs["desired_scope"])
        pts = int(15 * ratio)
        score += pts
        if matched:
            parts.append(f"dodatkowy zakres ({len(matched)}/{len(reqs['desired_scope'])})")

    # 6. OPGW (5 pts)
    if reqs["has_opgw"]:
        if _ref_has_scope_element(ref_scope, "opgw"):
            score += 5
            parts.append("OPGW w zakresie")
    else:
        score += 5

    # 7. Pole count (10 pts)
    if reqs["min_poles"]:
        if ref_poles is not None:
            if ref_poles >= reqs["min_poles"]:
                score += 10
                parts.append(f"słupy: {ref_poles} (≥{reqs['min_poles']})")
            elif ref_poles >= reqs["min_poles"] * 0.5:
                score += 5
                parts.append(f"słupy: {ref_poles} (<{reqs['min_poles']})")
            else:
                parts.append(f"słupy: {ref_poles} (znacznie poniżej {reqs['min_poles']})")
        else:
            parts.append(f"brak danych o ilości słupów")
    else:
        score += 10

    explanation = "; ".join(parts)
    return min(score, 100), explanation


# ---------------------------------------------------------------------------
# LLM explanations for top matches
# ---------------------------------------------------------------------------

def _generate_llm_explanations(
    top_matches: list[dict],
    tender_data: dict,
    reqs: dict,
) -> dict[str, str]:
    """Generate detailed Polish explanations for top 3 matches via Claude.

    Returns dict mapping filename -> explanation text.
    """
    try:
        import anthropic
    except ImportError:
        logger.warning("anthropic package not installed, skipping LLM explanations")
        return {}

    client = anthropic.Anthropic()

    tender_summary = (
        f"Przetarg: {tender_data.get('nazwa', '?')}\n"
        f"Zamawiający: {tender_data.get('zamawiajacy', '?')}\n"
        f"Zakres: {tender_data.get('zakres_prac', '?')}\n"
        f"Wymagania doświadczenia: {tender_data.get('wymagania_doswiadczenie', [])}\n"
    )

    results = {}
    for match in top_matches[:3]:
        ref_summary = json.dumps({
            "operator": match.get("operator"),
            "work_type": match.get("work_type"),
            "voltage": match.get("voltage"),
            "scope_elements": match.get("scope_elements"),
            "scope_summary": match.get("scope_summary"),
            "contract_value_pln_netto": match.get("contract_value_pln_netto"),
            "pole_count": match.get("pole_count"),
            "line_route": match.get("line_route"),
            "line_length_km": match.get("line_length_km"),
        }, ensure_ascii=False, indent=2)

        prompt = f"""Jesteś ekspertem od przetargów na roboty elektroenergetyczne.

PRZETARG (wymagania):
{tender_summary}

REFERENCJA (co wykonano):
{ref_summary}

Dopasowanie techniczne: {match['technical_score']}%

Napisz 3-5 zdań po polsku wyjaśniających DLACZEGO ta referencja pasuje (lub nie pasuje) do tego przetargu pod kątem TECHNICZNYM.
Porównaj:
1. Typ prac (przebudowa vs modernizacja vs wymiana OPGW)
2. Zakres elementów (słupy, fundamenty, przewody, OPGW, izolatory)
3. Skalę (ilość słupów, napięcie, długość linii)

Bądź konkretny, odnoś się do danych. Nie powtarzaj surowych liczb — interpretuj je.
Odpowiedz WYŁĄCZNIE tekstem wyjaśnienia, bez nagłówków."""

        try:
            response = client.messages.create(
                model="claude-sonnet-4-20250514",
                max_tokens=500,
                temperature=0.2,
                messages=[{"role": "user", "content": prompt}],
            )
            results[match["filename"]] = response.content[0].text.strip()
        except Exception as e:
            logger.warning(f"LLM explanation failed for {match['filename']}: {e}")
            results[match["filename"]] = match.get("technical_explanation", "")

    return results


# ---------------------------------------------------------------------------
# Main matching: evaluate ALL references on both axes
# ---------------------------------------------------------------------------

def match_all_references(tender_data: dict, references: list[dict]) -> list[dict]:
    """Evaluate every reference on two independent axes:
      1. Amount gate: TAK/NIE
      2. Technical match: 0-100%

    Returns ALL references with scores, sorted by technical_score desc.
    """
    reqs = _extract_tender_requirements(tender_data)
    results = []

    for ref in references:
        ref_name = ref.get("_source_file", ref.get("_db_file", "?"))

        # Gate 1: Amount
        amount_pass, amount_reason = _check_amount_gate(ref, reqs["min_amount"])

        # Gate 2: Technical score
        tech_score, tech_explanation = _score_technical_match(ref, reqs)

        results.append({
            "filename": ref_name,
            "operator": _normalize_operator(ref.get("operator", "")),
            "voltage": _normalize_voltage(ref.get("voltage")),
            "work_type": _normalize_work_type(ref.get("work_type")),
            "scope_elements": ref.get("scope_elements", []),
            "scope_summary": ref.get("scope_summary"),
            "contract_value_pln_netto": ref.get("contract_value_pln_netto"),
            "pole_count": ref.get("pole_count"),
            "line_route": ref.get("line_route"),
            "line_length_km": ref.get("line_length_km"),
            "amount_pass": amount_pass,
            "amount_reason": amount_reason,
            "technical_score": tech_score,
            "technical_explanation": tech_explanation,
            "llm_explanation": "",  # Filled later for top 3
        })

    results.sort(key=lambda x: x["technical_score"], reverse=True)
    return results


# ---------------------------------------------------------------------------
# XLSX report
# ---------------------------------------------------------------------------

def write_report_xlsx(
    all_matches: list[dict],
    tender_data: dict,
    output_path: Path,
    llm_explanations: dict[str, str] | None = None,
):
    """Write full comparison table with all references."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Analiza referencji"

    # Styles
    hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Calibri", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    orange_fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    reqs = _extract_tender_requirements(tender_data)

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"].value = f"Analiza referencji — {tender_data.get('nazwa', '')}"
    ws["A1"].font = Font(name="Calibri", size=12, bold=True)

    ws.merge_cells("A2:I2")
    ws["A2"].value = f"Wygenerowano: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="666666")

    # Requirements summary
    req_parts = []
    if reqs["work_type"]:
        req_parts.append(f"Typ: {reqs['work_type']}")
    if reqs["voltage"]:
        req_parts.append(f"Napięcie: ≥{reqs['voltage']}")
    if reqs["min_poles"]:
        req_parts.append(f"Słupy: ≥{reqs['min_poles']}")
    if reqs["required_scope"]:
        req_parts.append(f"Wymagany zakres: {', '.join(reqs['required_scope'])}")
    if reqs["min_amount"]:
        req_parts.append(f"Min. wartość: {reqs['min_amount']:,.0f} PLN")

    ws.merge_cells("A3:I3")
    ws["A3"].value = f"Wymagania przetargowe: {' | '.join(req_parts)}"
    ws["A3"].font = Font(name="Calibri", size=9, bold=True, color="333333")

    # Headers
    headers = [
        "Lp.", "Referencja", "Operator", "Typ prac",
        "Wartość (PLN netto)", "Kwota\n(TAK/NIE)", "Dopasowanie\ntechniczne",
        "Opis dopasowania", "Szczegółowa analiza (TOP 3)",
    ]
    col_widths = [5, 38, 12, 16, 18, 12, 14, 55, 55]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border
        # Column width via letter
        col_letter = chr(64 + col_idx) if col_idx <= 26 else chr(64 + col_idx // 26) + chr(64 + col_idx % 26)
        ws.column_dimensions[col_letter].width = width

    # Data rows
    for i, match in enumerate(all_matches, 1):
        row = i + 5

        # Lp.
        ws.cell(row=row, column=1, value=i).font = cell_font

        # Name
        name = match.get("line_route") or match.get("scope_summary") or match["filename"]
        if len(name) > 80:
            name = name[:77] + "..."
        ws.cell(row=row, column=2, value=name).font = cell_font

        # Operator
        ws.cell(row=row, column=3, value=match.get("operator", "—")).font = cell_font

        # Work type
        ws.cell(row=row, column=4, value=match.get("work_type", "—")).font = cell_font

        # Amount
        ref_val = match.get("contract_value_pln_netto")
        if ref_val is not None:
            c = ws.cell(row=row, column=5, value=ref_val)
            c.number_format = '#,##0'
        else:
            ws.cell(row=row, column=5, value="—")
        ws.cell(row=row, column=5).font = cell_font

        # Amount gate
        gate_cell = ws.cell(row=row, column=6)
        if match["amount_pass"]:
            gate_cell.value = "TAK"
            gate_cell.fill = pass_fill
        else:
            gate_cell.value = "NIE"
            gate_cell.fill = fail_fill
        gate_cell.font = Font(name="Calibri", size=10, bold=True)
        gate_cell.alignment = Alignment(horizontal="center")

        # Technical score
        tech = match["technical_score"]
        score_cell = ws.cell(row=row, column=7, value=f"{tech}%")
        score_cell.font = Font(name="Calibri", size=10, bold=True)
        score_cell.alignment = Alignment(horizontal="center")
        if tech >= 80:
            score_cell.fill = green_fill
        elif tech >= 60:
            score_cell.fill = yellow_fill
        elif tech >= 40:
            score_cell.fill = orange_fill
        else:
            score_cell.fill = red_fill

        # Short explanation
        ws.cell(row=row, column=8, value=match["technical_explanation"]).font = cell_font

        # LLM explanation (top 3 only)
        llm_text = ""
        if llm_explanations:
            llm_text = llm_explanations.get(match["filename"], "")
        ws.cell(row=row, column=9, value=llm_text).font = cell_font

        # Borders
        for col in range(1, 10):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            c.alignment = cell_align

    ws.freeze_panes = "A6"
    wb.save(output_path)
    logger.info(f"Report XLSX saved: {output_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def run_reference_matching(
    tender_data: dict,
    references_dir: Path,
    output_dir: Path,
    tender_hash: str | None = None,
    force: bool = False,
) -> Path | None:
    """Main entry point: load reference DB, match all, generate LLM explanations, write XLSX."""
    if not DB_DIR.exists() or not list(DB_DIR.glob("*.json")):
        logger.warning("No reference database found. Run extract_references.py first.")
        return None

    if tender_hash:
        output_dir = output_dir / tender_hash
        output_dir.mkdir(parents=True, exist_ok=True)

    # Skip-if-done
    existing = sorted(output_dir.glob("referencje_v*_*.xlsx"))
    if existing and not force:
        latest = existing[-1]
        db_files = sorted(DB_DIR.glob("*.json"))
        if db_files:
            newest_db = max(p.stat().st_mtime for p in db_files)
            if latest.stat().st_mtime > newest_db:
                logger.info(f"Skip-if-done: XLSX up to date ({latest.name})")
                return latest

    references = load_reference_db()
    if not references:
        return None

    # Match all
    all_matches = match_all_references(tender_data, references)
    above_50 = [m for m in all_matches if m["technical_score"] >= 50]
    logger.info(
        f"Reference matching: {len(all_matches)} total, "
        f"{len(above_50)} with technical ≥50%"
    )

    # LLM explanations for top 3
    top3 = all_matches[:3]
    llm_explanations = _generate_llm_explanations(top3, tender_data, _extract_tender_requirements(tender_data))

    # Write report
    version = len(existing) + 1
    date_str = datetime.now().strftime("%Y-%m-%d")
    xlsx_path = output_dir / f"referencje_v{version}_{date_str}.xlsx"
    write_report_xlsx(all_matches, tender_data, xlsx_path, llm_explanations)

    return xlsx_path
